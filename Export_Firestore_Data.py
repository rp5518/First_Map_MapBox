"""
Export_Firestore_Data.py

Downloads live canvassing marker states and notes from Firebase Firestore (maps4canvasing)
and merges them with household data from markers.json.

Exports to a formatted Excel file (Canvassing_Firestore_Report.xlsx) with auto-filters
and freeze panes for easy sorting and filtering.
"""

import json
import os
import re
import sys
import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuration
FIREBASE_PROJECT_ID = "maps4canvasing"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MARKERS_JSON_PATH = os.path.join(SCRIPT_DIR, "markers.json")
OUTPUT_EXCEL_PATH = os.path.join(SCRIPT_DIR, "Canvassing_Firestore_Report.xlsx")

def create_address_id(address):
    """Sanitizes address to match index.html createAddressId logic."""
    if not address:
        return ""
    return re.sub(r'[^a-zA-Z0-9]', '_', address).lower()

def parse_firestore_value(field_dict):
    """Recursively converts Firestore REST API field values into Python primitives."""
    if not isinstance(field_dict, dict) or not field_dict:
        return None
    if "stringValue" in field_dict:
        return field_dict["stringValue"]
    elif "booleanValue" in field_dict:
        return field_dict["booleanValue"]
    elif "integerValue" in field_dict:
        return int(field_dict["integerValue"])
    elif "doubleValue" in field_dict:
        return float(field_dict["doubleValue"])
    elif "timestampValue" in field_dict:
        return field_dict["timestampValue"]
    elif "mapValue" in field_dict:
        fields = field_dict.get("mapValue", {}).get("fields", {})
        return {k: parse_firestore_value(v) for k, v in fields.items()}
    elif "arrayValue" in field_dict:
        values = field_dict.get("arrayValue", {}).get("values", [])
        return [parse_firestore_value(v) for v in values]
    elif "nullValue" in field_dict:
        return None
    return None

def fetch_all_firestore_states(project_id):
    """Fetches all documents from the markerStates collection via Firestore REST API."""
    url = f"https://firestore.googleapis.com/v1/projects/{project_id}/databases/(default)/documents/markerStates"
    page_size = 300
    next_page_token = None
    all_docs = {}

    print(f"Fetching live marker states from Firestore project '{project_id}'...")
    while True:
        params = {"pageSize": page_size}
        if next_page_token:
            params["pageToken"] = next_page_token
        
        response = requests.get(url, params=params)
        if response.status_code != 200:
            print(f"Error fetching from Firestore: {response.status_code} {response.text}")
            break
        
        data = response.json()
        docs = data.get("documents", [])
        
        for doc in docs:
            doc_id = doc["name"].split("/")[-1]
            raw_fields = doc.get("fields", {})
            parsed_fields = {k: parse_firestore_value(v) for k, v in raw_fields.items()}
            all_docs[doc_id] = parsed_fields
            
        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break

    print(f"Retrieved {len(all_docs)} documents from Firestore.")
    return all_docs

def load_markers_json(json_path):
    """Loads markers.json if available."""
    if os.path.exists(json_path):
        print(f"Loading markers from: {json_path}")
        with open(json_path, "r", encoding="utf-8") as f:
            return json.load(f)
    else:
        print(f"Warning: {json_path} not found. Exporting Firestore data only.")
        return []

def process_data(markers, firestore_docs):
    """Combines markers.json and Firestore markerStates into export rows."""
    rows = []
    processed_doc_ids = set()

    # 1. Process markers from markers.json first (preserving map ordering)
    for idx, m in enumerate(markers, start=1):
        addr = m.get("address", "")
        addr_id = create_address_id(addr)
        processed_doc_ids.add(addr_id)
        
        f_doc = firestore_docs.get(addr_id, {})
        
        state_raw = f_doc.get("state", "")
        if state_raw == "visited":
            outcome = "Dropped Literature"
        elif state_raw == "friendly":
            outcome = "Friendly"
        elif state_raw == "notfriendly":
            outcome = "Not Friendly"
        else:
            outcome = "Unvisited" if not state_raw else state_raw

        flag = f_doc.get("flag", False)
        updated_by = f_doc.get("updatedBy", "")
        
        if updated_by == "proximity":
            method = "Auto (GPS Proximity)"
        elif state_raw != "":
            method = "Manual"
        else:
            method = "Unvisited"

        manual_loc = f_doc.get("manualUpdateLocation") or {}
        manual_dist = f_doc.get("manualUpdateDistanceFeet")
        manual_lat = manual_loc.get("lat") if isinstance(manual_loc, dict) else None
        manual_lng = manual_loc.get("lng") if isinstance(manual_loc, dict) else None
        manual_acc = manual_loc.get("accuracyFeet") if isinstance(manual_loc, dict) else None
        manual_time = manual_loc.get("timestamp") if isinstance(manual_loc, dict) else None

        notes_history = f_doc.get("notesHistory") or []
        formatted_notes = ""
        latest_note = ""
        if isinstance(notes_history, list) and len(notes_history) > 0:
            note_strings = []
            for n in notes_history:
                if isinstance(n, dict):
                    ts = n.get("timestamp", "")
                    txt = n.get("text", "")
                    note_strings.append(f"[{ts}] {txt}")
            formatted_notes = " | ".join(note_strings)
            latest_note = notes_history[-1].get("text", "") if isinstance(notes_history[-1], dict) else ""

        rows.append({
            "Map Index": idx,
            "Zone": m.get("zone", ""),
            "Address": addr,
            "First Name": m.get("first", ""),
            "Last Name": m.get("last", ""),
            "Party": m.get("party", ""),
            "Age": m.get("age", ""),
            "State / Outcome": outcome,
            "American Flag": "Yes" if flag else "No",
            "Update Method": method,
            "Manual Dist (ft)": manual_dist,
            "Manual GPS Lat": manual_lat,
            "Manual GPS Lng": manual_lng,
            "Manual GPS Accuracy (ft)": manual_acc,
            "Manual GPS Time": manual_time,
            "Proximity Timestamp": f_doc.get("proximityTimestamp", ""),
            "Latest Note": latest_note,
            "Full Notes History": formatted_notes,
            "Marker Lat": m.get("lat"),
            "Marker Lng": m.get("lng"),
            "Firestore Doc ID": addr_id
        })

    # 2. Process any Firestore docs that were NOT in markers.json
    extra_idx = len(markers) + 1
    for doc_id, f_doc in firestore_docs.items():
        if doc_id in processed_doc_ids:
            continue
        
        state_raw = f_doc.get("state", "")
        if state_raw == "visited":
            outcome = "Dropped Literature"
        elif state_raw == "friendly":
            outcome = "Friendly"
        elif state_raw == "notfriendly":
            outcome = "Not Friendly"
        else:
            outcome = "Unvisited" if not state_raw else state_raw

        flag = f_doc.get("flag", False)
        updated_by = f_doc.get("updatedBy", "")
        
        if updated_by == "proximity":
            method = "Auto (GPS Proximity)"
        elif state_raw != "":
            method = "Manual"
        else:
            method = "Unvisited"

        manual_loc = f_doc.get("manualUpdateLocation") or {}
        manual_dist = f_doc.get("manualUpdateDistanceFeet")
        manual_lat = manual_loc.get("lat") if isinstance(manual_loc, dict) else None
        manual_lng = manual_loc.get("lng") if isinstance(manual_loc, dict) else None
        manual_acc = manual_loc.get("accuracyFeet") if isinstance(manual_loc, dict) else None
        manual_time = manual_loc.get("timestamp") if isinstance(manual_loc, dict) else None

        notes_history = f_doc.get("notesHistory") or []
        formatted_notes = ""
        latest_note = ""
        if isinstance(notes_history, list) and len(notes_history) > 0:
            note_strings = []
            for n in notes_history:
                if isinstance(n, dict):
                    ts = n.get("timestamp", "")
                    txt = n.get("text", "")
                    note_strings.append(f"[{ts}] {txt}")
            formatted_notes = " | ".join(note_strings)
            latest_note = notes_history[-1].get("text", "") if isinstance(notes_history[-1], dict) else ""

        rows.append({
            "Map Index": extra_idx,
            "Zone": "Unknown",
            "Address": doc_id.replace("_", " ").title(),
            "First Name": "",
            "Last Name": "",
            "Party": "",
            "Age": "",
            "State / Outcome": outcome,
            "American Flag": "Yes" if flag else "No",
            "Update Method": method,
            "Manual Dist (ft)": manual_dist,
            "Manual GPS Lat": manual_lat,
            "Manual GPS Lng": manual_lng,
            "Manual GPS Accuracy (ft)": manual_acc,
            "Manual GPS Time": manual_time,
            "Proximity Timestamp": f_doc.get("proximityTimestamp", ""),
            "Latest Note": latest_note,
            "Full Notes History": formatted_notes,
            "Marker Lat": None,
            "Marker Lng": None,
            "Firestore Doc ID": doc_id
        })
        extra_idx += 1

    return rows

def format_excel(excel_path):
    """Formats the generated Excel workbook with headers, column widths, freeze panes, and auto-filters."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    ws.title = "Canvassing Report"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply auto filter & freeze header
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # Set row height
    ws.row_dimensions[1].height = 28

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            cell.border = thin_border
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Add padding
        adjusted_width = min(max(max_len + 3, 12), 60)
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(excel_path)
    print(f"Successfully formatted Excel report with Auto-Filters and Freeze Panes.")

def main():
    firestore_docs = fetch_all_firestore_states(FIREBASE_PROJECT_ID)
    markers = load_markers_json(MARKERS_JSON_PATH)
    
    data_rows = process_data(markers, firestore_docs)
    df = pd.DataFrame(data_rows)

    df.to_excel(OUTPUT_EXCEL_PATH, index=False)
    format_excel(OUTPUT_EXCEL_PATH)

    print("\n" + "="*50)
    print("EXPORT SUMMARY")
    print("="*50)
    print(f"Total Rows Exported: {len(df)}")
    if "State / Outcome" in df.columns:
        print("\nOutcome Breakdown:")
        print(df["State / Outcome"].value_counts().to_string())
    if "Update Method" in df.columns:
        print("\nUpdate Method Breakdown:")
        print(df["Update Method"].value_counts().to_string())
    if "American Flag" in df.columns:
        print("\nAmerican Flags:")
        print(df["American Flag"].value_counts().to_string())
    print("\nReport Saved To:")
    print(OUTPUT_EXCEL_PATH)
    print("="*50)

if __name__ == "__main__":
    main()
